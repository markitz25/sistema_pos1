"""
Vistas para el módulo de reportes
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q, Avg
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ventas.models import Venta, DetalleVenta
from inventario.models import Producto, Categoria
from clientes.models import Cliente
from trabajadores.models import Trabajador


def admin_required(function):
    """Decorador para verificar que el usuario sea administrador"""
    def wrap(request, *args, **kwargs):
        try:
            if request.user.trabajador.rol != 'admin':
                from django.contrib import messages
                messages.error(request, 'No tienes permisos para acceder a esta sección')
                return redirect('tablero:dashboard')
        except:
            from django.contrib import messages
            messages.error(request, 'Debes ser un trabajador registrado')
            return redirect('tablero:dashboard')
        return function(request, *args, **kwargs)
    return wrap


@login_required
@admin_required
def dashboard(request):
    """Dashboard principal de reportes"""
    
    # Fechas
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    # Estadísticas generales
    total_ventas_mes = Venta.objects.filter(
        fecha__gte=inicio_mes,
        estado='completada'
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    cantidad_ventas_mes = Venta.objects.filter(
        fecha__gte=inicio_mes,
        estado='completada'
    ).count()
    
    # Producto más vendido
    producto_top = DetalleVenta.objects.filter(
        venta__estado='completada'
    ).values('producto__nombre').annotate(
        total=Sum('cantidad')
    ).order_by('-total').first()
    
    # Cliente con más compras
    cliente_top = Cliente.objects.filter(
        activo=True
    ).order_by('-total_compras').first()
    
    context = {
        'total_ventas_mes': total_ventas_mes,
        'cantidad_ventas_mes': cantidad_ventas_mes,
        'producto_top': producto_top,
        'cliente_top': cliente_top,
    }
    
    return render(request, 'reportes/dashboard.html', context)


@login_required
@admin_required
def reporte_ventas(request):
    """Reporte detallado de ventas"""
    
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    metodo_pago = request.GET.get('metodo_pago')
    
    # Filtros base
    ventas = Venta.objects.filter(estado='completada')
    
    # Aplicar filtros
    if fecha_inicio:
        ventas = ventas.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        ventas = ventas.filter(fecha__lte=fecha_fin)
    if metodo_pago:
        ventas = ventas.filter(metodo_pago=metodo_pago)
    
    # Estadísticas
    total_ventas = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0')
    cantidad_ventas = ventas.count()
    promedio_venta = ventas.aggregate(promedio=Avg('total'))['promedio'] or Decimal('0')
    
    # Ventas por método de pago
    ventas_por_metodo = ventas.values('metodo_pago').annotate(
        total=Sum('total'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Últimas ventas
    ventas_recientes = ventas.select_related(
        'cliente', 'usuario'
    ).order_by('-fecha')[:20]
    
    context = {
        'ventas_recientes': ventas_recientes,
        'total_ventas': total_ventas,
        'cantidad_ventas': cantidad_ventas,
        'promedio_venta': promedio_venta,
        'ventas_por_metodo': ventas_por_metodo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'metodo_pago': metodo_pago,
    }
    
    return render(request, 'reportes/ventas.html', context)


@login_required
@admin_required
def reporte_productos(request):
    """Reporte de productos más vendidos"""
    
    # Productos más vendidos
    productos_vendidos = DetalleVenta.objects.filter(
        venta__estado='completada'
    ).values(
        'producto__id',
        'producto__nombre',
        'producto__categoria__nombre',
        'producto__precio'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        ingresos_generados=Sum(F('cantidad') * F('precio_unitario'))
    ).order_by('-cantidad_vendida')[:20]
    
    # Productos con stock bajo
    productos_stock_bajo = Producto.objects.filter(
        stock__lt=10
    ).select_related('categoria').order_by('stock')[:20]
    
    # Estadísticas por categoría
    ventas_por_categoria = DetalleVenta.objects.filter(
        venta__estado='completada'
    ).values(
        'producto__categoria__nombre'
    ).annotate(
        total_ventas=Sum(F('cantidad') * F('precio_unitario')),
        unidades_vendidas=Sum('cantidad')
    ).order_by('-total_ventas')
    
    context = {
        'productos_vendidos': productos_vendidos,
        'productos_stock_bajo': productos_stock_bajo,
        'ventas_por_categoria': ventas_por_categoria,
    }
    
    return render(request, 'reportes/productos.html', context)


@login_required
@admin_required
def reporte_clientes(request):
    """Reporte de clientes"""
    
    # Top clientes
    top_clientes = Cliente.objects.filter(
        activo=True
    ).annotate(
        total_compras_count=Count('ventas')
    ).order_by('-total_compras')[:20]
    
    # Clientes nuevos este mes
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    clientes_nuevos = Cliente.objects.filter(
        creado_en__gte=inicio_mes
    ).count()
    
    # Estadísticas generales
    total_clientes = Cliente.objects.filter(activo=True).count()
    
    context = {
        'top_clientes': top_clientes,
        'clientes_nuevos': clientes_nuevos,
        'total_clientes': total_clientes,
    }
    
    return render(request, 'reportes/clientes.html', context)


@login_required
@admin_required
def exportar_excel(request):
    """Exportar reporte de ventas a Excel"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ['ID', 'Fecha', 'Cliente', 'Subtotal', 'IVA', 'Descuento', 'Total', 'Método Pago', 'Estado']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    ventas = Venta.objects.select_related('cliente').order_by('-fecha')[:100]
    
    for row, venta in enumerate(ventas, start=2):
        ws.cell(row=row, column=1, value=venta.id)
        ws.cell(row=row, column=2, value=venta.fecha.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=venta.cliente.nombre)
        ws.cell(row=row, column=4, value=float(venta.subtotal))
        ws.cell(row=row, column=5, value=float(venta.impuesto))
        ws.cell(row=row, column=6, value=float(venta.descuento))
        ws.cell(row=row, column=7, value=float(venta.total))
        ws.cell(row=row, column=8, value=venta.get_metodo_pago_display())
        ws.cell(row=row, column=9, value=venta.get_estado_display())
    
    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Crear response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_ventas_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response